"""
APPeki - Calcolatore Iperbilirubinemia Neonatale
Basato sul foglio Excel _APPeki FINAL.xlsx e sul nomogramma di Bhutani.
"""

import streamlit as st
import openpyxl
import numpy as np
import plotly.graph_objects as go
from datetime import datetime, date, time
import bisect

# ─────────────────────────────────────────────────────────────────────────────
# Caricamento tabelle dall'Excel
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data
def load_tables():
    wb = openpyxl.load_workbook("_APPeki FINAL.xlsx", data_only=True)
    ws = wb["Foglio3"]

    # ── Tabella Bilirubinemia Totale Sierica (BTS) ───────────────────────────
    # Rows 149-221, cols B(ora), C(p50), D(p75), E(p90)
    bts_hours, bts_p50, bts_p75, bts_p90 = [], [], [], []
    for r in range(149, 222):
        h = ws.cell(r, 2).value
        if isinstance(h, (int, float)):
            bts_hours.append(float(h))
            bts_p50.append(ws.cell(r, 3).value)
            bts_p75.append(ws.cell(r, 4).value)
            bts_p90.append(ws.cell(r, 5).value)

    # ── Tabella Bilirubinemia Transcutanea (BTC) ──────────────────────────────
    # Rows 228-300, cols B(ora), C(p50), D(p75)
    btc_hours, btc_p50, btc_p75 = [], [], []
    for r in range(228, 301):
        h = ws.cell(r, 2).value
        if isinstance(h, (int, float)):
            btc_hours.append(float(h))
            btc_p50.append(ws.cell(r, 3).value)
            btc_p75.append(ws.cell(r, 4).value)

    # ── Tabella Soglia Fototerapia (FT) e Exsanguinotrasfusione (EXT) ─────────
    # Rows 149-269, cols J(ora), K-O(FT per EG), Q-V(EXT per EG)
    ft_hours, ft_lt30, ft_30_32, ft_32_35, ft_35_38, ft_gte38 = [], [], [], [], [], []
    ext_hours, ext_lt30, ext_lt32, ext_lt35, ext_lt38, ext_gte38 = [], [], [], [], [], []
    for r in range(149, 270):
        h_ft = ws.cell(r, 10).value   # col J
        h_ext = ws.cell(r, 17).value  # col Q
        if isinstance(h_ft, (int, float)):
            ft_hours.append(float(h_ft))
            ft_lt30.append(ws.cell(r, 11).value)   # K: EG<30
            ft_30_32.append(ws.cell(r, 12).value)  # L: EG 30-31.6
            ft_32_35.append(ws.cell(r, 13).value)  # M: EG 32-34.6
            ft_35_38.append(ws.cell(r, 14).value)  # N: EG 35-37.6
            ft_gte38.append(ws.cell(r, 15).value)  # O: EG>=38
        if isinstance(h_ext, (int, float)):
            ext_hours.append(float(h_ext))
            ext_lt30.append(ws.cell(r, 18).value)  # R
            ext_lt32.append(ws.cell(r, 19).value)  # S
            ext_lt35.append(ws.cell(r, 20).value)  # T
            ext_lt38.append(ws.cell(r, 21).value)  # U
            ext_gte38.append(ws.cell(r, 22).value) # V

    return {
        "bts": {"hours": bts_hours, "p50": bts_p50, "p75": bts_p75, "p90": bts_p90},
        "btc": {"hours": btc_hours, "p50": btc_p50, "p75": btc_p75},
        "ft":  {"hours": ft_hours,  "lt30": ft_lt30, "3032": ft_30_32,
                "3235": ft_32_35,   "3538": ft_35_38, "gte38": ft_gte38},
        "ext": {"hours": ext_hours, "lt30": ext_lt30, "lt32": ext_lt32,
                "lt35": ext_lt35,   "lt38": ext_lt38, "gte38": ext_gte38},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Dati nomogramma di Bhutani (Bhutani VK, Pediatrics 1999;103:6-14)
# Percentili 40°, 75°, 95° della bilirubinemia sierica totale (mg/dL)
# per ore di vita postnatale
# ─────────────────────────────────────────────────────────────────────────────

BHUTANI_HOURS = [12, 18, 24, 30, 36, 42, 48, 54, 60, 66, 72, 78, 84, 90, 96,
                 102, 108, 114, 120, 126, 132, 138, 144]
BHUTANI_P40  = [3.0, 4.5, 6.0, 7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 9.8, 10.0,
                10.5, 11.0, 11.5, 12.0, 12.3, 12.6, 12.9, 13.0, 13.0, 13.0, 13.0, 13.0]
BHUTANI_P75  = [6.0, 7.5, 9.0, 10.5, 11.5, 12.0, 12.5, 13.0, 13.5, 14.0, 14.5,
                15.0, 15.0, 15.5, 15.5, 15.8, 16.0, 16.0, 16.0, 16.0, 16.0, 16.0, 16.0]
BHUTANI_P95  = [7.5, 9.5, 11.5, 12.5, 13.5, 14.5, 15.0, 15.5, 16.0, 16.5, 17.0,
                17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0, 17.0]


# ─────────────────────────────────────────────────────────────────────────────
# Funzioni di calcolo
# ─────────────────────────────────────────────────────────────────────────────

def match_approx(value: float, sorted_list: list) -> int:
    """
    Replica MATCH(value, sorted_list, 1):
    restituisce l'indice (0-based) del più grande elemento ≤ value.
    """
    idx = bisect.bisect_right(sorted_list, value) - 1
    return max(0, min(idx, len(sorted_list) - 1))


def lookup_bts(ore: float, tables: dict):
    """Restituisce (p50, p75, p90) per la BTS all'ora indicata."""
    t = tables["bts"]
    if ore < t["hours"][0]:
        return None, None, None
    i = match_approx(ore, t["hours"])
    return t["p50"][i], t["p75"][i], t["p90"][i]


def classify_bilirubin(value: float, p50: float, p75: float) -> str:
    if value is None or p50 is None or p75 is None:
        return "—"
    if value < p50:
        return "< 50°"
    elif value == p50:
        return "50°"
    elif value < p75:
        return "50°–75°"
    elif value == p75:
        return "75°"
    else:
        return "> 75°"


def lookup_btc(ore: float, tables: dict):
    """Restituisce (p50, p75) per la BTC all'ora indicata."""
    t = tables["btc"]
    if ore < t["hours"][0]:
        return None, None
    i = match_approx(ore, t["hours"])
    return t["p50"][i], t["p75"][i]


def get_ft_threshold(ore: float, eg: float, tables: dict) -> float | None:
    """Soglia di fototerapia (mg/dL) in base alle ore di vita e all'EG."""
    t = tables["ft"]
    if not t["hours"]:
        return None
    i = match_approx(ore, t["hours"])
    if eg < 30:
        return t["lt30"][i]
    elif eg < 32:
        return t["3032"][i]
    elif eg < 35:
        return t["3235"][i]
    elif eg < 38:
        return t["3538"][i]
    else:
        return t["gte38"][i]


def get_ext_threshold(ore: float, eg: float, tables: dict) -> float | None:
    """Soglia di exsanguinotrasfusione (mg/dL) in base alle ore di vita e all'EG."""
    t = tables["ext"]
    if not t["hours"]:
        return None
    i = match_approx(ore, t["hours"])
    if eg < 30:
        return t["lt30"][i]
    elif eg < 32:
        return t["lt32"][i]
    elif eg < 35:
        return t["lt35"][i]
    elif eg < 38:
        return t["lt38"][i]
    else:
        return t["gte38"][i]


def interp_bhutani(ore: float, curve: list) -> float:
    """Interpolazione lineare sul nomogramma di Bhutani."""
    if ore <= BHUTANI_HOURS[0]:
        return curve[0]
    if ore >= BHUTANI_HOURS[-1]:
        return curve[-1]
    idx = bisect.bisect_right(BHUTANI_HOURS, ore) - 1
    h0, h1 = BHUTANI_HOURS[idx], BHUTANI_HOURS[idx + 1]
    v0, v1 = curve[idx], curve[idx + 1]
    return v0 + (v1 - v0) * (ore - h0) / (h1 - h0)


def bhutani_zone(bts_value: float, ore: float) -> tuple[str, str]:
    """
    Restituisce (zona, colore) secondo il nomogramma di Bhutani.
    Applicabile per EG ≥ 35 settimane e peso ≥ 2000 g.
    """
    p40 = interp_bhutani(ore, BHUTANI_P40)
    p75 = interp_bhutani(ore, BHUTANI_P75)
    p95 = interp_bhutani(ore, BHUTANI_P95)
    if bts_value < p40:
        return "Low-risk zone", "#2ecc71"
    elif bts_value < p75:
        return "Low-intermediate-risk zone", "#f1c40f"
    elif bts_value < p95:
        return "High-intermediate-risk zone", "#e67e22"
    else:
        return "High-risk zone", "#e74c3c"


# ─────────────────────────────────────────────────────────────────────────────
# Costruzione grafico nomogramma di Bhutani con posizione paziente
# ─────────────────────────────────────────────────────────────────────────────

def build_bhutani_chart(ore_paziente: float | None = None,
                        bts_paziente: float | None = None):
    h = BHUTANI_HOURS
    p40 = BHUTANI_P40
    p75 = BHUTANI_P75
    p95 = BHUTANI_P95

    fig = go.Figure()

    # Riempimenti delle zone (dall'alto verso il basso)
    fig.add_trace(go.Scatter(
        x=h + h[::-1], y=p95 + [25] * len(h),
        fill="toself", fillcolor="rgba(231,76,60,0.15)",
        line=dict(color="rgba(0,0,0,0)"),
        name="High-risk zone", showlegend=True,
        hoverinfo="skip"
    ))
    fig.add_trace(go.Scatter(
        x=h + h[::-1], y=p75 + p95[::-1],
        fill="toself", fillcolor="rgba(230,126,34,0.15)",
        line=dict(color="rgba(0,0,0,0)"),
        name="High-intermediate-risk zone", showlegend=True,
        hoverinfo="skip"
    ))
    fig.add_trace(go.Scatter(
        x=h + h[::-1], y=p40 + p75[::-1],
        fill="toself", fillcolor="rgba(241,196,15,0.15)",
        line=dict(color="rgba(0,0,0,0)"),
        name="Low-intermediate-risk zone", showlegend=True,
        hoverinfo="skip"
    ))
    fig.add_trace(go.Scatter(
        x=h + h[::-1], y=[0] * len(h) + p40[::-1],
        fill="toself", fillcolor="rgba(46,204,113,0.15)",
        line=dict(color="rgba(0,0,0,0)"),
        name="Low-risk zone", showlegend=True,
        hoverinfo="skip"
    ))

    # Linee dei percentili
    fig.add_trace(go.Scatter(x=h, y=p95, mode="lines",
                             line=dict(color="#e74c3c", width=2),
                             name="95° percentile"))
    fig.add_trace(go.Scatter(x=h, y=p75, mode="lines",
                             line=dict(color="#e67e22", width=2),
                             name="75° percentile"))
    fig.add_trace(go.Scatter(x=h, y=p40, mode="lines",
                             line=dict(color="#2ecc71", width=2),
                             name="40° percentile"))

    # Punto paziente
    if ore_paziente is not None and bts_paziente is not None:
        zona, colore = bhutani_zone(bts_paziente, ore_paziente)
        fig.add_trace(go.Scatter(
            x=[ore_paziente], y=[bts_paziente],
            mode="markers",
            marker=dict(size=14, color=colore, symbol="star",
                        line=dict(color="black", width=1.5)),
            name=f"Paziente ({zona})",
        ))

    fig.update_layout(
        title="Nomogramma di Bhutani – Classificazione del rischio",
        xaxis=dict(title="Ore di vita postnatale", range=[0, 150],
                   dtick=12, gridcolor="#eeeeee"),
        yaxis=dict(title="Bilirubinemia totale sierica (mg/dL)",
                   range=[0, 25], gridcolor="#eeeeee"),
        legend=dict(x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.8)"),
        plot_bgcolor="white",
        height=520,
    )
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# Layout Streamlit
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="APPeki", page_icon="🍋", layout="wide")
st.title("🍋 APPeki – Calcolatore Iperbilirubinemia Neonatale")

tables = load_tables()

# ── Sezione input ─────────────────────────────────────────────────────────────
st.header("📋 Dati del paziente")

col1, col2 = st.columns(2)

with col1:
    st.subheader("Nascita")
    data_nascita = st.date_input("Data di nascita", value=date.today(),
                                 key="dn")
    ora_nascita  = st.time_input("Ora di nascita", value=time(8, 0),
                                 key="on", step=60)

with col2:
    st.subheader("Prelievo")
    data_prelievo = st.date_input("Data del prelievo", value=date.today(),
                                  key="dp")
    ora_prelievo  = st.time_input("Ora del prelievo", value=time(12, 0),
                                  key="op", step=60)

dt_nascita  = datetime.combine(data_nascita,  ora_nascita)
dt_prelievo = datetime.combine(data_prelievo, ora_prelievo)
delta_ore   = (dt_prelievo - dt_nascita).total_seconds() / 3600.0
ore_vita    = round(delta_ore)

if ore_vita < 0:
    st.error("⚠️  La data/ora del prelievo è precedente alla nascita.")
    st.stop()

st.divider()

col3, col4 = st.columns(2)

with col3:
    eg = st.number_input(
        "Età gestazionale (settimane)",
        min_value=22, max_value=44, value=38, step=1,
        help="Settimane complete di gestazione"
    )
    bts_value = st.number_input(
        "BTS – Bilirubinemia Totale Sierica (mg/dL)",
        min_value=0.0, max_value=30.0, value=0.0, step=0.1,
        help="Lasciare a 0 se non disponibile"
    )
    btc_value = st.number_input(
        "BTC – Bilirubinemia Transcutanea (mg/dL)",
        min_value=0.0, max_value=25.0, value=0.0, step=0.1,
        help="Lasciare a 0 se non disponibile"
    )

with col4:
    st.subheader("Fattori di rischio (Maisels et al. 2009)")
    rf_emolisi    = st.checkbox("Malattia emolitica iso-immune / G6PD / sfero")
    rf_allattamento = st.checkbox("Allattamento esclusivo con scarsa alimentazione")
    rf_ittero24   = st.checkbox("Ittero clinico nelle prime 24 ore")
    rf_fratello   = st.checkbox("Fratello/sorella con ittero grave trattato")
    rf_ecchimosi  = st.checkbox("Cefalematoma o ecchimosi significative")
    rf_asiatico   = st.checkbox("Origine asiatica orientale")
    rf_sepsi      = st.checkbox("Sepsi / acidosi")
    rf_albumina   = st.checkbox("Albumina < 3.0 g/dL")

n_fattori_rischio = sum([rf_emolisi, rf_allattamento, rf_ittero24,
                         rf_fratello, rf_ecchimosi, rf_asiatico,
                         rf_sepsi, rf_albumina])

# ── Calcoli ───────────────────────────────────────────────────────────────────
st.divider()
st.header("📊 Risultati")

# Ore di vita
col_a, col_b, col_c = st.columns(3)
with col_a:
    st.metric("Ore di vita", f"{ore_vita} h")
with col_b:
    prematurita = "PREMATURO" if eg < 37 else "A termine"
    st.metric("Maturità", prematurita)
with col_c:
    st.metric("Fattori di rischio", n_fattori_rischio)

st.divider()

# ── BTS ───────────────────────────────────────────────────────────────────────
st.subheader("🔬 Bilirubinemia Totale Sierica (BTS)")

bts_p50, bts_p75, bts_p90 = lookup_bts(ore_vita, tables)

if bts_p50 is None:
    st.info(f"Tabella BTS non disponibile per ore di vita < 24 h (ore attuali: {ore_vita} h).")
else:
    col1, col2, col3 = st.columns(3)
    col1.metric("50° percentile", f"{bts_p50:.1f} mg/dL")
    col2.metric("75° percentile", f"{bts_p75:.1f} mg/dL")
    col3.metric("90° percentile", f"{bts_p90:.1f} mg/dL")

    if bts_value > 0:
        perc_class = classify_bilirubin(bts_value, bts_p50, bts_p75)
        above_90 = bts_value > bts_p90
        flag_90 = " ⚠️  SOPRA IL 90°!" if above_90 else ""
        st.info(f"**BTS paziente ({bts_value:.1f} mg/dL)** → Percentile: **{perc_class}**{flag_90}")
        if eg < 37:
            st.warning("⚠️  PREMATURO – la classificazione percentile potrebbe non essere applicabile.")
    else:
        st.caption("(BTS non inserita)")

st.divider()

# ── BTC ───────────────────────────────────────────────────────────────────────
st.subheader("🔆 Bilirubinemia Transcutanea (BTC)")

btc_p50, btc_p75 = lookup_btc(ore_vita, tables)

if btc_p50 is None:
    st.info(f"Tabella BTC non disponibile per ore di vita < 25 h (ore attuali: {ore_vita} h).")
else:
    col1, col2 = st.columns(2)
    col1.metric("50° percentile", f"{btc_p50:.1f} mg/dL")
    col2.metric("75° percentile", f"{btc_p75:.1f} mg/dL")

    if btc_value > 0:
        perc_class_btc = classify_bilirubin(btc_value, btc_p50, btc_p75)
        st.info(f"**BTC paziente ({btc_value:.1f} mg/dL)** → Percentile: **{perc_class_btc}**")
        if eg < 37:
            st.warning("⚠️  PREMATURO – la classificazione percentile potrebbe non essere applicabile.")
    else:
        st.caption("(BTC non inserita)")

st.divider()

# ── Soglie terapeutiche ───────────────────────────────────────────────────────
st.subheader("💡 Soglie terapeutiche")

ft_soglia  = get_ft_threshold(ore_vita, eg, tables)
ext_soglia = get_ext_threshold(ore_vita, eg, tables)

col1, col2 = st.columns(2)

with col1:
    if ft_soglia is not None:
        st.metric("Soglia Fototerapia (FT)", f"{ft_soglia:.1f} mg/dL")
        if bts_value > 0 and bts_value >= ft_soglia:
            st.error(f"🔴 BTS ≥ soglia FT ({bts_value:.1f} ≥ {ft_soglia:.1f} mg/dL) → **Considerare fototerapia**")
        elif bts_value > 0:
            st.success(f"🟢 BTS sotto la soglia FT ({bts_value:.1f} < {ft_soglia:.1f} mg/dL)")
    else:
        st.caption("Soglia FT non disponibile")

with col2:
    if ext_soglia is not None:
        st.metric("Soglia Exsanguinotrasfusione (EXT)", f"{ext_soglia:.1f} mg/dL")
        if bts_value > 0 and bts_value >= ext_soglia:
            st.error(f"🔴 BTS ≥ soglia EXT ({bts_value:.1f} ≥ {ext_soglia:.1f} mg/dL) → **Considerare exsanguinotrasfusione**")
        elif bts_value > 0:
            st.success(f"🟢 BTS sotto la soglia EXT ({bts_value:.1f} < {ext_soglia:.1f} mg/dL)")
    else:
        st.caption("Soglia EXT non disponibile")

st.divider()

# ── Nomogramma di Bhutani ─────────────────────────────────────────────────────
st.subheader("📈 Nomogramma di Bhutani – Classe di rischio")

bhutani_applicable = eg >= 35
if not bhutani_applicable:
    st.warning("⚠️  Il nomogramma di Bhutani è applicabile a EG ≥ 35 settimane. "
               "Il grafico viene mostrato a titolo orientativo.")

# Calcola zona se BTS disponibile
if bts_value > 0:
    zona, colore = bhutani_zone(bts_value, float(ore_vita))

    # Badge zona
    zona_it = {
        "Low-risk zone": "Zona a basso rischio",
        "Low-intermediate-risk zone": "Zona a rischio intermedio-basso",
        "High-intermediate-risk zone": "Zona a rischio intermedio-alto",
        "High-risk zone": "Zona ad alto rischio",
    }.get(zona, zona)

    zona_desc = {
        "Low-risk zone": "Il paziente si trova nella zona a basso rischio di iperbilirubinemia severa.",
        "Low-intermediate-risk zone": ("Il paziente è nella zona a rischio intermedio-basso. "
                                       "Monitoraggio attento raccomandato."),
        "High-intermediate-risk zone": ("⚠️  Rischio intermedio-alto. Valutare i fattori di rischio aggiuntivi "
                                        "per decidere il follow-up."),
        "High-risk zone": ("🔴 ALTA PRIORITÀ. Il paziente è nella zona ad alto rischio. "
                           "Follow-up precoce e valutazione terapeutica raccomandata."),
    }.get(zona, "")

    st.markdown(
        f"<div style='background:{colore};padding:12px 20px;border-radius:8px;"
        f"color:white;font-size:1.2rem;font-weight:bold;margin-bottom:8px'>"
        f"🔷 {zona_it}</div>",
        unsafe_allow_html=True
    )
    st.markdown(zona_desc)

    if n_fattori_rischio > 0 and zona in ("Low-intermediate-risk zone",):
        st.info(f"ℹ️  Con {n_fattori_rischio} fattore/i di rischio presente/i, "
                "considerare un follow-up anticipato.")

chart = build_bhutani_chart(
    ore_paziente=float(ore_vita) if bts_value > 0 else None,
    bts_paziente=bts_value if bts_value > 0 else None,
)
st.plotly_chart(chart, use_container_width=True)

st.caption(
    "Fonte: Bhutani VK, Johnson L, Sivieri EM. Predictive ability of a predischarge "
    "hour-specific serum bilirubin for subsequent significant hyperbilirubinemia in "
    "healthy term and near-term newborns. Pediatrics. 1999;103:6-14. "
    "| Maisels MJ et al. Pediatrics. 2009;124:1193-1198."
)

# ── Riepilogo Fattori di Rischio ──────────────────────────────────────────────
if n_fattori_rischio > 0:
    st.divider()
    st.subheader("⚠️  Fattori di rischio presenti")
    fattori = {
        "Malattia emolitica iso-immune / G6PD / sferocitosi": rf_emolisi,
        "Allattamento esclusivo con scarsa alimentazione": rf_allattamento,
        "Ittero clinico nelle prime 24 ore": rf_ittero24,
        "Fratello/sorella con ittero grave trattato": rf_fratello,
        "Cefalematoma o ecchimosi significative": rf_ecchimosi,
        "Origine asiatica orientale": rf_asiatico,
        "Sepsi / acidosi": rf_sepsi,
        "Albumina < 3.0 g/dL": rf_albumina,
    }
    for nome, presente in fattori.items():
        if presente:
            st.markdown(f"- ✅ {nome}")
